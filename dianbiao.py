import os
import sys
import datetime
import json
import shutil
import time

# ==========================================
# 【字体配置】智能查找并注册中文字体
# ==========================================
from kivy.core.text import LabelBase


def get_chinese_font_path():
    if sys.platform == 'win32':
        fonts_dir = r'C:\Windows\Fonts'
        candidates = ['msyh.ttc', 'msyh.ttf', 'simhei.ttf', 'simsun.ttc']
        for font_file in candidates:
            full_path = os.path.join(fonts_dir, font_file)
            if os.path.exists(full_path):
                return full_path
        return 'msyh.ttc'
    elif sys.platform == 'android':
        return '/system/fonts/DroidSansFallbackFull.ttf'
    elif sys.platform == 'darwin':
        return '/System/Library/Fonts/PingFang.ttc'
    return None


font_path = get_chinese_font_path()
if font_path and os.path.exists(font_path):
    try:
        LabelBase.register(name='Roboto', fn_regular=font_path)
    except Exception:
        pass

# ==========================================

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment  # 新增：用于设置 Excel 样式
from kivy.graphics import Color, Rectangle

Window.clearcolor = (0.95, 0.95, 0.95, 1)


class SimpleMeterApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # 文件名定义
        self.main_xlsx = 'dianbiao.xlsx'  # 综合数据总表
        self.daily_xlsx = 'meiribiao.xlsx'  # 单日采集模板
        self.config_file = 'dianbiao.json'  # 统一配置文件

        # 确定存储路径
        if sys.platform == 'android':
            from android.storage import primary_external_storage_path
            self.base_path = primary_external_storage_path()
        else:
            self.base_path = os.getcwd()

        self.full_main_path = os.path.join(self.base_path, self.main_xlsx)
        self.full_daily_path = os.path.join(self.base_path, self.daily_xlsx)
        self.full_json_path = os.path.join(self.base_path, self.config_file)

        self.meter_list = []
        self.location_map = {}

        now = datetime.datetime.now()
        self.current_day = now.day
        self.current_month = now.month
        self.current_year = now.year

        self.load_config()
        self.load_meter_list_from_excel()

        # 初始化单日表（智能判断是否需要重建）
        self.init_daily_excel()

    def load_config(self):
        """从 dianbiao.json 加载位置信息"""
        if not os.path.exists(self.full_json_path):
            self.location_map = {}
            return
        try:
            with open(self.full_json_path, 'r', encoding='utf-8') as f:
                self.location_map = json.load(f)
        except Exception:
            self.location_map = {}

    def load_meter_list_from_excel(self):
        """从综合表 dianbiao.xlsx 读取表号列表"""
        if not os.path.exists(self.full_main_path):
            self.meter_list = [f'表{i}' for i in range(1, 19)]
            return

        try:
            wb = load_workbook(self.full_main_path, data_only=True)
            ws = wb.active
            temp_list = []
            for r in range(3, 50):
                val = ws.cell(row=r, column=1).value
                if val is not None and str(val).strip() != "":
                    temp_list.append(str(val).strip())
            wb.close()
            self.meter_list = temp_list if temp_list else [f'表{i}' for i in range(1, 19)]
        except Exception:
            self.meter_list = [f'表{i}' for i in range(1, 19)]

    def init_daily_excel(self):
        """
        智能初始化单日采集模板 meiribiao.xlsx
        逻辑：
        1. 如果文件不存在 -> 创建新表。
        2. 如果文件存在 -> 检查标题行日期。
           - 日期是今天 -> 不做任何操作（保留当天已录数据）。
           - 日期不是今天 -> 删除旧文件，创建新表（自动重建）。
        """
        today_str = f"{self.current_year}年{self.current_month}月{self.current_day}日 抄表记录"

        need_rebuild = False

        # 1. 检查文件是否存在
        if os.path.exists(self.full_daily_path):
            try:
                # 尝试读取现有文件的日期
                wb_check = load_workbook(self.full_daily_path, data_only=True)
                ws_check = wb_check.active
                existing_title = ws_check.cell(row=1, column=1).value
                wb_check.close()

                # 2. 对比日期
                if existing_title == today_str:
                    print(f"每日表日期正确 ({today_str})，保留现有数据，不重建。")
                    return  # 日期一致，直接退出，不执行后续重建逻辑
                else:
                    print(f"检测到旧日期 ({existing_title})，需要重建新表。")
                    need_rebuild = True
            except Exception as e:
                # 如果文件损坏或读取失败，也视为需要重建
                print(f"读取现有每日表失败 ({e})，将重建新表。")
                need_rebuild = True
        else:
            print("每日表不存在，将创建新表。")
            need_rebuild = True

        # 3. 执行重建逻辑 (如果需要)
        if need_rebuild:
            # 如果文件存在但日期不对，先删除
            if os.path.exists(self.full_daily_path):
                try:
                    os.remove(self.full_daily_path)
                    print(f"已删除旧的每日表：{self.full_daily_path}")
                except Exception as e:
                    print(f"删除旧每日表失败：{e}")

            # 创建全新的空白表
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "每日采集"

                # --- 匹配图片格式 ---
                # 第1行：合并单元格显示大标题日期
                ws.merge_cells('A1:B1')
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = today_str
                title_cell.font = Font(bold=True, size=16)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 第2行：表头 (表号，度数)
                header_font = Font(bold=True)
                h1 = ws.cell(row=2, column=1, value="表号")
                h1.font = header_font
                h2 = ws.cell(row=2, column=2, value="度数")
                h2.font = header_font

                # 第3行开始：预填表号 (对应图片中的表1-表18)
                for i, meter in enumerate(self.meter_list):
                    row = i + 3
                    ws.cell(row=row, column=1).value = meter
                    # 图片中没有位置列，所以这里不填位置

                # 调整列宽
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 20

                wb.save(self.full_daily_path)
                wb.close()
                print(f"已创建新的每日表：{self.full_daily_path}")
            except Exception as e:
                print(f"初始化单日表失败：{e}")

    def build(self):
        layout = BoxLayout(orientation='vertical', padding=20, spacing=10)

        # 1. 标题区
        main_title = Label(
            text="每日电量抄录系统",
            size_hint=(1, 0.12),
            font_size='24sp',
            bold=True,
            halign='center',
            valign='middle',
            color=(0, 0, 0, 1)
        )
        main_title.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
        layout.add_widget(main_title)

        date_label = Label(
            text=f"当前日期：{self.current_year}-{self.current_month:02d}-{self.current_day:02d}",
            size_hint=(1, 0.08),
            font_size='16sp',
            color=(0.2, 0.5, 0.8, 1),
            bold=True,
            halign='center',
            valign='middle'
        )
        layout.add_widget(date_label)

        spacer = Label(size_hint=(1, 0.05))
        layout.add_widget(spacer)

        # 2. 控件区
        self.meter_spinner = Spinner(
            text='请选择表号',
            values=self.meter_list,
            size_hint=(1, 0.12),
            font_size='18sp',
            background_color=(1, 1, 1, 1),
            color=(0, 0, 0, 1)
        )
        self.meter_spinner.bind(text=self.on_meter_select)
        layout.add_widget(self.meter_spinner)

        self.location_label = Label(
            text="位置：等待选择...",
            size_hint=(1, 0.08),
            font_size='14sp',
            color=(0.3, 0.3, 0.3, 1),
            halign='left',
            valign='middle',
            padding=(10, 0)
        )
        self.location_label.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
        layout.add_widget(self.location_label)

        self.reading_input = TextInput(
            hint_text='输入当前读数 (数字)',
            multiline=False,
            input_filter='float',
            size_hint=(1, 0.15),
            font_size='22sp',
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            cursor_color=(0, 0, 0, 1)
        )
        self.reading_input.bind(on_text_validate=self.save_data)
        layout.add_widget(self.reading_input)

        self.status_label = Label(
            text="就绪 - 请选择表号并录入",
            size_hint=(1, 0.1),
            font_size='14sp',
            color=(0.5, 0.5, 0.5, 1)
        )
        layout.add_widget(self.status_label)

        # 3. 按钮区
        btn_layout = BoxLayout(size_hint=(1, 0.12), spacing=5)

        save_btn = Button(
            text="保存",
            font_size='13sp',
            background_color=(0, 0.8, 0, 0.9),
            color=(1, 1, 1, 1)
        )
        save_btn.bind(on_press=self.save_data)

        view_btn = Button(
            text="查看今日",
            font_size='13sp',
            background_color=(0.6, 0.6, 0.6, 0.9),
            color=(1, 1, 1, 1)
        )
        view_btn.bind(on_press=self.show_today_data)

        share_btn = Button(
            text="分享",
            font_size='13sp',
            background_color=(0, 0.5, 1, 0.9),
            color=(1, 1, 1, 1)
        )
        share_btn.bind(on_press=self.show_share_options)

        import_btn = Button(
            text="导入/同步",
            font_size='13sp',
            background_color=(0.8, 0.5, 0, 0.9),
            color=(1, 1, 1, 1)
        )
        import_btn.bind(on_press=self.start_import_process)

        btn_layout.add_widget(save_btn)
        btn_layout.add_widget(view_btn)
        btn_layout.add_widget(share_btn)
        btn_layout.add_widget(import_btn)

        layout.add_widget(btn_layout)

        return layout

    def on_meter_select(self, instance, value):
        if value == '请选择表号':
            self.location_label.text = "位置：等待选择..."
            self.location_label.color = (0.3, 0.3, 0.3, 1)
            self.reading_input.text = ''
            self.show_status("就绪", (0.5, 0.5, 0.5, 1))
            return
        location_info = self.location_map.get(value, "位置未配置")
        self.location_label.text = f"位置：{location_info}"
        self.location_label.color = (0, 0.5, 0.8, 1)
        self.reading_input.text = ''
        self.reading_input.focus = True
        self.show_status(f"已选择：{value}", (0, 0.5, 1, 1))

    def get_yesterday_reading(self, meter_name):
        """获取该表号昨天的读数，用于校验"""
        if not os.path.exists(self.full_main_path):
            return None

        try:
            wb = load_workbook(self.full_main_path, data_only=True)
            ws = wb.active

            # 1. 找到表号行
            target_row = None
            for r in range(3, 50):
                val = ws.cell(row=r, column=1).value
                if val is not None and str(val).strip() == meter_name:
                    target_row = r
                    break

            if target_row is None:
                wb.close()
                return None

            # 2. 找到昨天的日期列
            yesterday = self.current_day - 1
            if yesterday <= 0:
                wb.close()
                return None

            target_col = None
            for c in range(2, 100):
                h_val = ws.cell(row=2, column=c).value
                if h_val is not None:
                    try:
                        if int(float(h_val)) == yesterday:
                            target_col = c
                            break
                    except:
                        continue

            if target_col:
                val = ws.cell(row=target_row, column=target_col).value
                wb.close()
                return val if val is not None else None

            wb.close()
            return None
        except Exception:
            return None

    def save_data(self, instance):
        meter_name = self.meter_spinner.text
        reading_str = self.reading_input.text.strip()

        if meter_name == '请选择表号' or not reading_str:
            self.show_status("请选择表号并输入读数", (1, 0, 0, 1))
            return

        try:
            reading = float(reading_str)

            # === 数据校验逻辑 ===
            yesterday_val = self.get_yesterday_reading(meter_name)
            if yesterday_val is not None:
                if reading < yesterday_val:
                    self.show_error_popup(meter_name, reading, yesterday_val)
                    return

                    # 1. 写入单日表 (meiribiao.xlsx)
            self.save_to_daily_excel(meter_name, reading)

            # 2. 写入综合表 (dianbiao.xlsx)
            self.save_to_main_excel(meter_name, self.current_day, reading)

            self.show_status("保存成功! (已同步至总表)", (0, 0.8, 0, 1))
            self.reading_input.text = ''
            self.meter_spinner.text = '请选择表号'
            self.location_label.text = "位置：等待选择..."
            self.location_label.color = (0.3, 0.3, 0.3, 1)
            self.reading_input.focus = False

            # 刷新单日表中的显示（更新日期标题）
            self.init_daily_excel()

        except Exception as e:
            self.show_status(f"失败：{str(e)}", (1, 0, 0, 1))

    def show_error_popup(self, meter_name, current_val, last_val):
        """显示错误警告弹窗"""
        content = BoxLayout(orientation='vertical', padding=20, spacing=15)

        lbl = Label(
            text=f"数据异常提醒\n\n表号：{meter_name}\n昨日读数：{last_val}\n今日输入：{current_val}\n\n今日读数不能小于昨日读数！\n请核对是否抄错或电表归零。",
            halign='center',
            valign='middle',
            font_size='16sp',
            color=(1, 0, 0, 1)
        )
        lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))

        btn_close = Button(
            text="返回修改",
            size_hint_y=None,
            height=50,
            background_color=(1, 0.5, 0, 1),
            color=(1, 1, 1, 1)
        )

        popup = Popup(title="逻辑错误", content=content, size_hint=(0.85, 0.5), auto_dismiss=False)
        btn_close.bind(on_press=lambda x: popup.dismiss())

        content.add_widget(lbl)
        content.add_widget(btn_close)
        popup.open()

    def save_to_daily_excel(self, meter_name, reading):
        """保存到单日采集表 meiribiao.xlsx (适配新格式：数据从第3行开始)"""
        if not os.path.exists(self.full_daily_path):
            self.init_daily_excel()

        wb = load_workbook(self.full_daily_path)
        ws = wb.active

        # 更新标题日期
        ws.cell(row=1, column=1).value = f"{self.current_year}年{self.current_month}月{self.current_day}日 抄表记录"

        # 查找表号行并更新 (数据从第3行开始，因为第1行标题，第2行表头)
        found = False
        for r in range(3, 50):
            m_val = ws.cell(row=r, column=1).value
            if m_val and str(m_val).strip() == meter_name:
                ws.cell(row=r, column=2).value = reading  # 写入 B 列 (度数)
                found = True
                break

        if not found:
            # 如果是新表号，追加到最后
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1).value = meter_name
            ws.cell(row=next_row, column=2).value = reading

        wb.save(self.full_daily_path)
        wb.close()

    def save_to_main_excel(self, meter_name, day, reading):
        """保存到综合数据表 dianbiao.xlsx"""
        if not os.path.exists(self.full_main_path):
            self.create_main_excel_structure()

        wb = load_workbook(self.full_main_path)
        ws = wb.active

        # 1. 找行
        target_row = None
        for r in range(3, 50):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() == meter_name:
                target_row = r
                break

        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1).value = meter_name

        # 2. 找列 (日期)
        target_col = None
        for c in range(2, 100):
            h_val = ws.cell(row=2, column=c).value
            if h_val is not None:
                try:
                    if int(float(h_val)) == day:
                        target_col = c
                        break
                except:
                    continue

        if target_col is None:
            target_col = ws.max_column + 1
            ws.cell(row=2, column=target_col).value = day

        # 3. 写入数据
        ws.cell(row=target_row, column=target_col).value = reading
        wb.save(self.full_main_path)
        wb.close()

    def create_main_excel_structure(self):
        """创建综合表的基础结构"""
        wb = Workbook()
        ws = wb.active
        ws.title = "综合数据"
        ws.cell(row=1, column=1).value = "表号"
        ws.cell(row=2, column=1).value = "日期"

        for i, meter in enumerate(self.meter_list):
            ws.cell(row=i + 3, column=1).value = meter

        wb.save(self.full_main_path)
        wb.close()

    def show_today_data(self, instance):
        """查看今日数据 (优先读单日表，如果没有则读综合表)"""
        source_file = self.full_daily_path
        source_name = "单日采集表"

        use_main = False
        if not os.path.exists(source_file):
            use_main = True
        else:
            try:
                wb = load_workbook(source_file, data_only=True)
                ws = wb.active
                has_data = False
                # 检查数据区域 (从第3行开始)
                for r in range(3, 50):
                    if ws.cell(row=r, column=2).value is not None:
                        has_data = True
                        break
                wb.close()
                if not has_data:
                    use_main = True
            except:
                use_main = True

        if use_main:
            source_file = self.full_main_path
            source_name = "综合数据表"
            self._show_data_from_file(source_file, source_name, use_main_flag=True)
        else:
            self._show_data_from_file(source_file, source_name, use_main_flag=False)

    def _show_data_from_file(self, file_path, source_name, use_main_flag):
        if not os.path.exists(file_path):
            self.show_status("数据文件不存在", (1, 0, 0, 1))
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            content_layout = BoxLayout(orientation='vertical', padding=10, spacing=5)
            with content_layout.canvas.before:
                Color(1, 1, 1, 1)
                self.popup_bg_rect = Rectangle(pos=content_layout.pos, size=content_layout.size)
            content_layout.bind(
                pos=lambda inst, val: setattr(self, 'popup_bg_rect', self.popup_bg_rect) or setattr(self.popup_bg_rect,
                                                                                                    'pos', val),
                size=lambda inst, val: setattr(self, 'popup_bg_rect', self.popup_bg_rect) or setattr(self.popup_bg_rect,
                                                                                                     'size', val))

            scroll_view = ScrollView(size_hint=(1, 1))
            data_layout = BoxLayout(orientation='vertical', size_hint_y=None, padding=10, spacing=10)
            data_layout.bind(minimum_height=data_layout.setter('height'))

            with data_layout.canvas.before:
                Color(1, 1, 1, 1)
                self.data_bg_rect = Rectangle(pos=data_layout.pos, size=data_layout.size)
            data_layout.bind(
                pos=lambda inst, val: setattr(self, 'data_bg_rect', self.data_bg_rect) or setattr(self.data_bg_rect,
                                                                                                  'pos', val),
                size=lambda inst, val: setattr(self, 'data_bg_rect', self.data_bg_rect) or setattr(self.data_bg_rect,
                                                                                                   'size', val))

            count = 0

            if use_main_flag:
                # 综合表逻辑：找今日列
                target_col = None
                for c in range(2, 100):
                    h_val = ws.cell(row=2, column=c).value
                    if h_val is not None:
                        try:
                            if int(float(h_val)) == self.current_day:
                                target_col = c
                                break
                        except:
                            continue

                if target_col:
                    for r in range(3, 50):
                        meter_name = ws.cell(row=r, column=1).value
                        val = ws.cell(row=r, column=target_col).value
                        if meter_name and val is not None:
                            count += 1
                            row_text = f"{meter_name}: {val}"
                            item_label = Label(text=row_text, size_hint_y=None, height=40, halign='left',
                                               valign='middle', font_size='16sp', color=(0, 0, 0, 1), padding=(10, 0))
                            item_label.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
                            data_layout.add_widget(item_label)
                else:
                    lbl = Label(text=f"{source_name} 中今日 ({self.current_day}日) 暂无数据。", size_hint_y=None,
                                height=100, halign='center', valign='middle', font_size='16sp',
                                color=(0.5, 0.5, 0.5, 1))
                    data_layout.add_widget(lbl)
            else:
                # 单日表逻辑：直接读 B 列 (度数)，从第3行开始
                for r in range(3, 50):
                    meter_name = ws.cell(row=r, column=1).value
                    val = ws.cell(row=r, column=2).value
                    if meter_name and val is not None:
                        count += 1
                        row_text = f"{meter_name}: {val}"
                        item_label = Label(text=row_text, size_hint_y=None, height=40, halign='left', valign='middle',
                                           font_size='16sp', color=(0, 0, 0, 1), padding=(10, 0))
                        item_label.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
                        data_layout.add_widget(item_label)

                if count == 0:
                    lbl = Label(text=f"{source_name} 中暂无数据。", size_hint_y=None, height=100, halign='center',
                                valign='middle', font_size='16sp', color=(0.5, 0.5, 0.5, 1))
                    data_layout.add_widget(lbl)

            scroll_view.add_widget(data_layout)
            content_layout.add_widget(scroll_view)
            wb.close()

            close_btn = Button(text="关闭", size_hint_y=None, height=50, background_color=(0.5, 0.5, 0.5, 1),
                               color=(1, 1, 1, 1))
            title_text = f"今日数据 ({source_name})"
            popup = Popup(title=title_text, content=content_layout, size_hint=(0.9, 0.7), auto_dismiss=False)
            close_btn.bind(on_press=lambda x: popup.dismiss())
            content_layout.add_widget(close_btn)
            popup.open()

        except Exception as e:
            self.show_status(f"读取失败：{str(e)}", (1, 0, 0, 1))

    def show_share_options(self, instance):
        """显示分享选项弹窗"""
        content = BoxLayout(orientation='vertical', spacing=15, padding=20)
        lbl = Label(
            text="请选择要分享的文件：",
            halign='center',
            valign='middle',
            font_size='18sp',
            size_hint_y=None,
            height=50
        )

        btn_daily = Button(
            text="分享单日数据 (meiribiao.xlsx)",
            size_hint_y=None,
            height=60,
            background_color=(0, 0.6, 0.8, 1),
            color=(1, 1, 1, 1)
        )

        btn_main = Button(
            text="分享综合数据 (dianbiao.xlsx)",
            size_hint_y=None,
            height=60,
            background_color=(0.8, 0.5, 0, 1),
            color=(1, 1, 1, 1)
        )

        btn_cancel = Button(text="取消", size_hint_y=None, height=50)

        popup = Popup(title="分享选择", content=content, size_hint=(0.85, 0.5), auto_dismiss=False)

        def do_share(file_type):
            popup.dismiss()
            if file_type == 'daily':
                self.share_file(self.full_daily_path, "每日电量采集表")
            else:
                self.share_file(self.full_main_path, "综合电量数据表")

        btn_daily.bind(on_press=lambda x: do_share('daily'))
        btn_main.bind(on_press=lambda x: do_share('main'))
        btn_cancel.bind(on_press=lambda x: popup.dismiss())

        content.add_widget(lbl)
        content.add_widget(btn_daily)
        content.add_widget(btn_main)
        content.add_widget(btn_cancel)
        popup.open()

    def share_file(self, file_path, subject_prefix):
        if not os.path.exists(file_path):
            self.show_status("文件不存在", (1, 0, 0, 1))
            return
        try:
            from android import mActivity
            from jnius import autoclass
            Intent = autoclass('android.content.Intent')
            File = autoclass('java.io.File')
            Uri = autoclass('android.net.Uri')

            intent = Intent(Intent.ACTION_SEND)
            file_obj = File(file_path)
            uri = Uri.fromFile(file_obj)

            intent.setType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            intent.putExtra(Intent.EXTRA_STREAM, uri)
            intent.putExtra(Intent.EXTRA_SUBJECT, f"{subject_prefix}-{self.current_year}{self.current_month:02d}")

            mActivity.startActivity(Intent.createChooser(intent, "分享文件"))
        except ImportError:
            self.show_status(f"文件路径：{file_path}", (0, 0.5, 1, 1))
            print(f"File path: {file_path}")

    def start_import_process(self, instance):
        self.show_import_instructions()

    def show_import_instructions(self):
        content = BoxLayout(orientation='vertical', spacing=15, padding=20)
        lbl = Label(
            text="【数据同步与迁移】\n\n\n\n\n\n"
                 "1. 更新配置/表号：\n"
                 "   发送 update_dianbiao.xlsx 和 update_dianbiao.json 到手机根目录。\n\n"
                 "2. 合并历史数据：\n"
                 "   如果在其他设备录入了 meiribiao.xlsx，可将其重命名为 update_meiribiao.xlsx 放入根目录进行合并。\n\n"
                 "点击下方按钮开始操作。",
            size_hint_y=None,
            height=240,
            halign='center',
            valign='middle',
            font_size='13sp'
        )
        lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))

        btn_config = Button(text="1. 更新配置与表号", size_hint_y=None, height=50, background_color=(0, 0.6, 0.8, 1),
                            color=(1, 1, 1, 1))
        btn_merge = Button(text="2. 合并单日数据", size_hint_y=None, height=50, background_color=(0.8, 0.5, 0, 1),
                           color=(1, 1, 1, 1))
        btn_close = Button(text="取消", size_hint_y=None, height=50)

        popup = Popup(title="导入向导", content=content, size_hint=(0.9, 0.7), auto_dismiss=False)

        def select_action(action):
            popup.dismiss()
            if action == 'config':
                self.perform_config_migration()
            elif action == 'merge':
                self.merge_daily_data()

        btn_config.bind(on_press=lambda x: select_action('config'))
        btn_merge.bind(on_press=lambda x: select_action('merge'))
        btn_close.bind(on_press=lambda x: popup.dismiss())

        content.add_widget(lbl)
        content.add_widget(btn_config)
        content.add_widget(btn_merge)
        content.add_widget(btn_close)
        popup.open()

    def perform_config_migration(self):
        """更新配置和主表结构"""
        base = self.base_path
        temp_xlsx = os.path.join(base, "update_dianbiao.xlsx")
        temp_json = os.path.join(base, "update_dianbiao.json")

        updated = False

        if os.path.exists(temp_json):
            try:
                shutil.copy(temp_json, self.full_json_path)
                self.load_config()
                os.remove(temp_json)
                updated = True
            except Exception as e:
                self.show_status(f"配置更新失败：{e}", (1, 0, 0, 1))
                return

        if os.path.exists(temp_xlsx):
            try:
                self.migrate_excel_data(temp_xlsx, self.full_main_path)
                os.remove(temp_xlsx)
                updated = True
            except Exception as e:
                self.show_status(f"表格更新失败：{e}", (1, 0, 0, 1))
                return

        if updated:
            self.load_meter_list_from_excel()
            self.show_restart_popup()
        else:
            self.show_status("未找到更新文件 (update_dianbiao.xlsx/json)", (1, 0, 0, 1))

    def merge_daily_data(self):
        """将 update_meiribiao.xlsx 的数据合并到主表"""
        base = self.base_path
        temp_daily = os.path.join(base, "update_meiribiao.xlsx")

        if not os.path.exists(temp_daily):
            self.show_status("未找到 update_meiribiao.xlsx", (1, 0, 0, 1))
            return

        try:
            wb_temp = load_workbook(temp_daily, data_only=True)
            ws_temp = wb_temp.active

            if not os.path.exists(self.full_main_path):
                self.create_main_excel_structure()

            wb_main = load_workbook(self.full_main_path)
            ws_main = wb_main.active

            count = 0
            # 遍历临时文件的每一行 (从第3行开始，因为第1行标题，第2行表头)
            for r in range(3, 100):
                meter = ws_temp.cell(row=r, column=1).value
                val = ws_temp.cell(row=r, column=2).value  # 读取 B 列

                if meter and val is not None:
                    target_row = None
                    for mr in range(3, 100):
                        if str(ws_main.cell(row=mr, column=1).value) == str(meter):
                            target_row = mr
                            break

                    if target_row is None:
                        target_row = ws_main.max_row + 1
                        ws_main.cell(row=target_row, column=1).value = meter

                    target_col = None
                    for c in range(2, 100):
                        h_val = ws_main.cell(row=2, column=c).value
                        if h_val is not None and int(float(h_val)) == self.current_day:
                            target_col = c
                            break

                    if target_col is None:
                        target_col = ws_main.max_column + 1
                        ws_main.cell(row=2, column=target_col).value = self.current_day

                    ws_main.cell(row=target_row, column=target_col).value = val
                    count += 1

            wb_main.save(self.full_main_path)
            wb_main.close()
            wb_temp.close()

            os.remove(temp_daily)
            self.show_status(f"成功合并 {count} 条数据到总表", (0, 0.8, 0, 1))
            self.show_restart_popup()

        except Exception as e:
            self.show_status(f"合并失败：{e}", (1, 0, 0, 1))

    def migrate_excel_data(self, new_file_path, old_file_path):
        """合并主表数据逻辑"""
        if not os.path.exists(old_file_path):
            shutil.copy(new_file_path, old_file_path)
            return

        wb_new = load_workbook(new_file_path)
        ws_new = wb_new.active
        wb_old = load_workbook(old_file_path)
        ws_old = wb_old.active

        old_data_map = {}
        for r in range(3, 100):
            meter_name = ws_old.cell(row=r, column=1).value
            if meter_name:
                meter_name = str(meter_name).strip()
                old_data_map[meter_name] = {}
                for c in range(2, 100):
                    date_val = ws_old.cell(row=2, column=c).value
                    cell_val = ws_old.cell(row=r, column=c).value
                    if date_val is not None and cell_val is not None:
                        try:
                            d_key = int(float(date_val))
                            old_data_map[meter_name][d_key] = cell_val
                        except:
                            continue

        for r in range(3, 100):
            meter_name = ws_new.cell(row=r, column=1).value
            if meter_name:
                meter_name = str(meter_name).strip()
                if meter_name in old_data_map:
                    row_data = old_data_map[meter_name]
                    for c in range(2, 100):
                        date_val = ws_new.cell(row=2, column=c).value
                        if date_val is not None:
                            try:
                                d_key = int(float(date_val))
                                if d_key in row_data:
                                    ws_new.cell(row=r, column=c).value = row_data[d_key]
                            except:
                                continue

        wb_new.save(old_file_path)
        wb_new.close()
        wb_old.close()

    def show_restart_popup(self):
        content = Label(text="操作成功！\n\n请重启 APP 以生效。", halign='center', valign='middle', font_size='18sp')
        content.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
        popup = Popup(title="完成", content=content, size_hint=(0.8, 0.4), auto_dismiss=False)
        btn_restart = Button(text="立即重启", size_hint_y=None, height=60, background_color=(0, 0.5, 1, 1),
                             color=(1, 1, 1, 1))

        def restart_app(*args):
            popup.dismiss()
            if sys.platform == 'android':
                from android import mActivity
                mActivity.recreate()
            else:
                App.get_running_app().stop()

        btn_restart.bind(on_press=restart_app)
        box = BoxLayout(orientation='vertical')
        box.add_widget(content)
        box.add_widget(btn_restart)
        popup.content = box
        popup.open()

    def show_status(self, text, color):
        self.status_label.text = text
        self.status_label.color = color


if __name__ == '__main__':
    SimpleMeterApp().run()